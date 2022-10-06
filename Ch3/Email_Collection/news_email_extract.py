import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook


# url = "https://n.news.naver.com/mnews/article/081/0003305201?sid=105"
url = "https://n.news.naver.com/mnews/article/018/0005329984?sid=105"

headers = {
    'User-Agent': 'Mozilla/5.0',
    'Content-Type': 'text/html; charset=utf-8'
}

response = requests.get(url, headers=headers)

results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)
results = list(set(results))

print(results)

try:
    wb = load_workbook(r'email.xlsx', data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active

for result in results:
    sheet.append([result])

wb.save('email.xlsx')