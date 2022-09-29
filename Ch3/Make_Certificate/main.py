from openpyxl import load_workbook
from make_Certification import Certification
from convert_PDF import Convert2PDF

load_wb = load_workbook(r'certification.xlsx')
load_ws = load_wb.active

name_list = []
birthday_list = []
ho_list = []

for i in range(1, load_ws.max_row + 1):
    name_list.append(load_ws.cell(i,1).value)
    birthday_list.append(load_ws.cell(i,2).value)
    ho_list.append(load_ws.cell(i,3).value)

print(name_list)
print(birthday_list)
print(ho_list)

def make_docx():
    for i in range(0, len(name_list)):
        c = Certification(name_list[i], birthday_list[i], ho_list[i])
        c.make()

def make_pdf():
    for i in range(0, len(name_list)):
        c = Convert2PDF(name_list[i], birthday_list[i], ho_list[i])
        c.convert()


make_pdf()